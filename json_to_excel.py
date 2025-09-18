
"""
Convertit un dossier de JSON (une fiche par fichier OU liste de fiches) en un Excel.
- Pour chaque JSON, récupère: `title`, `content`, et **tous les champs** présents
  dans `fields` (insensible à la casse: accepte `fields`, `Fields`, etc.).
- **N'UTILISE PAS** `field_schema` / `Field_schema`.
- Ajoute une colonne `id_dossier` (1..N) numérotée **dans l'ordre des fichiers** (tri lexicographique)
  et, pour les JSON-listes, dans l'ordre d'apparition.
- **Sanitise** toutes les chaînes pour Excel (suppression des caractères illégaux openpyxl)
  et **tronque** au besoin (par défaut 32 760 caractères) pour éviter les erreurs.

Usage:
  python3 json_to_excel.py \
    --src "/Users/ethanrouzier/Documents/Fiches acces precoce" \
    --out "/Users/ethanrouzier/Documents/Fiches_acces_precoce.xlsx" \
    --sheet "export" \
    --max-cell-chars 32760

Dépendances:
  pip install pandas openpyxl
"""
from __future__ import annotations
import argparse
import json
from pathlib import Path
from typing import Dict, Any, Iterable, List
import pandas as pd
import re

# openpyxl fournit une regex des caractères illégaux pour Excel
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as _XL_BAD
except Exception:
    _XL_BAD = re.compile(r"[---]")

FIELDS_KEYS = {"fields", "Fields", "FIELDS", "field", "Field"}
SCHEMA_KEYS = {"field_schema", "Field_schema", "FIELDS_SCHEMA", "schema", "Schema"}

import re
from typing import Any

# Caractères interdits en XML 1.0 (sauf \t \n \r)
_XML_BAD = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

_FORMULA_TRIGGERS = ("=", "+", "-", "@")  # début de chaîne dangereux pour Excel

def sanitize_excel_value(val: Any, max_len: int = 32760) -> str:
    """
    - Convertit en str
    - Retire caractères XML interdits
    - Neutralise un début de chaîne qui déclencherait une formule Excel
    - Tronque à max_len (limite Excel ≈ 32767)
    """
    s = "" if val is None else str(val)

    # Normaliser fin de ligne
    s = s.replace("\r\n", "\n").replace("\r", "\n")

    # Retirer caractères XML interdits
    s = _XML_BAD.sub("", s)

    # Neutraliser formules Excel
    if s and s[0] in _FORMULA_TRIGGERS:
        # Prefixer par apostrophe → Excel l’interprète comme du texte
        s = "'" + s

    # Tronquage (en gardant une marque)
    if max_len and len(s) > max_len:
        s = s[: max_len - 10] + " [trunc]"

    return s


def _read_json(path: Path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _get_fields(d: Dict[str, Any]) -> Dict[str, Any]:
    for k in FIELDS_KEYS:
        v = d.get(k)
        if isinstance(v, dict):
            return v
    return {}


def _clean_for_excel(s: str, max_len: int | None = 32760) -> str:
    if not isinstance(s, str):
        s = str(s)
    s = s.replace("", "")  # null explicite si présent
    s = _XL_BAD.sub("", s)
    if max_len and len(s) > max_len:
        s = s[:max_len]
    return s


def _normalize_scalar(v: Any, max_len: int | None) -> Any:
    if v is None:
        return ""
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        return _clean_for_excel(v, max_len)
    # dict/list/objet complexe → string JSON lisible, puis nettoyage
    try:
        return _clean_for_excel(json.dumps(v, ensure_ascii=False), max_len)
    except Exception:
        return _clean_for_excel(str(v), max_len)


def _iter_docs_from_file(path: Path) -> Iterable[Dict[str, Any]]:
    data = _read_json(path)
    if isinstance(data, dict):
        yield data
    elif isinstance(data, list):
        for item in data:
            if isinstance(item, dict):
                yield item


from pathlib import Path
from typing import List, Dict, Any

# suppose que sanitize_excel_value, _iter_docs_from_file, _get_fields, _normalize_scalar existent

def collect_rows(src: Path, max_len: int | None) -> List[Dict[str, Any]]:
    """
    Parcourt tous les JSON sous 'src' et construit les lignes d'export Excel.
    - title/content/category toujours présents
    - date_doc pris depuis les fields (ou doc) si dispo
    - anti-formule (= + - @ en début), nettoyage XML, tronquage (max_len)
    """
    rows: List[Dict[str, Any]] = []
    max_len = max_len or 32760

    files = sorted([p for p in src.rglob("*.json") if p.is_file()])
    for f in files:
        for doc in _iter_docs_from_file(f):
            # Base
            row: Dict[str, Any] = {
                "title":    sanitize_excel_value(doc.get("title", ""), max_len),
                "content":  sanitize_excel_value(doc.get("content", ""), max_len),
                "category": sanitize_excel_value(doc.get("category_name") or doc.get("category") or "", max_len),
            }

            # Champs extraits
            fields = _get_fields(doc) or {}
            for k, v in fields.items():
                row[str(k)] = sanitize_excel_value(_normalize_scalar(v, max_len), max_len)

            # date_doc si disponible (depuis fields ou doc)
            if "date_doc" not in row or not row["date_doc"]:
                if "date" in fields and fields["date"]:
                    row["date_doc"] = sanitize_excel_value(fields["date"], max_len)
                elif doc.get("date"):
                    row["date_doc"] = sanitize_excel_value(doc.get("date"), max_len)

            # Ceinture + bretelles : (re)sanitize tout
            for k in list(row.keys()):
                row[k] = sanitize_excel_value(row[k], max_len)

            rows.append(row)

    return rows


def to_excel(rows: list[dict], out_path: Path, sheet: str, max_len: int):
    import pandas as pd

    if not rows:
        # écrire quand même un fichier minimal pour éviter les surprises
        df = pd.DataFrame([{"title": "", "content": "", "category": "", "date_doc": ""}])
    else:
        df = pd.DataFrame(rows)
    df.insert(0, "id_dossier", range(2, len(df) + 2))  
    # Forcer type objet et sanitize partout
    for c in df.columns:
        df[c] = df[c].astype("object").map(lambda x: sanitize_excel_value(x, max_len))

    # Écriture
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as ew:
        df.to_excel(ew, index=False, sheet_name=sheet)
    with pd.ExcelWriter(out_path, engine="openpyxl") as ew:
        df.to_excel(ew, index=False, sheet_name="export")
        wb = ew.book
        # retire les autres feuilles vides (souvent la sheet par défaut)
        for name in list(wb.sheetnames):
            if name == "export":
                continue
            ws = wb[name]
            if ws.max_row <= 1 and ws.max_column <= 1 and (ws["A1"].value in (None, "")):
                wb.remove(ws)


def main():
    ap = argparse.ArgumentParser(description="Export JSON -> Excel (title, content, category, fields)")
    ap.add_argument("--src", required=True, type=Path, help="Dossier racine des JSON")
    ap.add_argument("--out", required=True, type=Path, help="Chemin du fichier .xlsx de sortie")
    ap.add_argument("--sheet", default="export", help="Nom de l'onglet Excel")
    ap.add_argument("--max-cell-chars", type=int, default=32760, help="Taille max par cellule (Excel ≈ 32767)")
    args = ap.parse_args()

    rows = collect_rows(args.src, args.max_cell_chars)
    to_excel(rows, args.out, args.sheet, args.max_cell_chars)
    print(f"✅ Exporté {len(rows)} lignes vers {args.out} (onglet: {args.sheet})")


if __name__ == "__main__":
    main()

